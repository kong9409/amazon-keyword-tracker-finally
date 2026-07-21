from __future__ import annotations

import argparse
import json
import mimetypes
import os
import re
import shutil
import sqlite3
import threading
import time
import uuid
import webbrowser
from datetime import datetime, timedelta, timezone
from email.parser import BytesParser
from email.policy import default as email_default_policy
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse
import ipaddress
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from cryptography.fernet import Fernet, InvalidToken
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from lark_writer import append_records_to_lark
from sorftime_adapter import build_sorftime_client, test_sorftime_connection

BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"

# One web application for Zeabur or local use. Zeabur injects PORT automatically.
APP_MODE = (os.getenv("APP_MODE") or ("hosted" if os.getenv("PORT") else "local")).strip().lower()
HOSTED_MODE = APP_MODE in {"hosted", "zeabur", "cloud"}
APP_HOST = os.getenv("HOST") or ("0.0.0.0" if HOSTED_MODE else "127.0.0.1")
APP_PORT = int(os.getenv("PORT") or os.getenv("KEYWORD_TRACKER_PORT", "8766"))
DEFAULT_TIMEZONE = os.getenv("APP_TIMEZONE", "Asia/Shanghai")

# Zeabur should mount a persistent volume at /app/data. Daily configuration,
# encrypted credentials, history and scheduled Excel files all live there.
DATA_DIR = Path(os.getenv("APP_DATA_DIR") or (BASE_DIR / "data")).expanduser()
EXPORT_DIR = Path(os.getenv("APP_EXPORT_DIR") or (DATA_DIR / "exports" if HOSTED_MODE else BASE_DIR / "exports")).expanduser()
JOB_DIR = DATA_DIR / "jobs"
DAILY_DIR = DATA_DIR / "daily_jobs"
CONNECTION_DIR = DATA_DIR / "local_connections"
DB_PATH = DATA_DIR / "keyword_tracker.db"
PID_PATH = DATA_DIR / "app.pid"
SCHEDULER_KEY_PATH = DATA_DIR / ".scheduler.key"
HOSTED_MCP_HOSTS = [item.strip().lower() for item in os.getenv("SORFTIME_ALLOWED_MCP_HOSTS", "").split(",") if item.strip()]
EXCEL_FONT = "Microsoft YaHei"

# Required operational fields are deliberately first and fully visible in UI/Excel.
FIELD_COLUMNS: list[tuple[str, str]] = [
    ("date", "日期"),
    ("asin", "ASIN"),
    ("keyword", "关键词"),
    ("traffic_share", "流量占比"),
    ("aba_rank", "ABA热度"),
    ("search_volume", "搜索量"),
    ("organic_position", "自然位"),
    ("ad_position", "广告位"),
    ("price", "价格"),
    ("coupon_value", "优惠券"),
    ("deal_price", "秒杀价"),
    ("prime_discount_price", "Prime价"),
    ("estimated_sales", "月销量"),
    ("product_rank", "大类排名"),
    ("rating", "评分"),
    ("review_count", "评价数"),
    ("product_url", "链接"),
    ("captured_at", "抓取时间"),
    ("marketplace", "站点"),
    ("organic_time", "自然曝光时间"),
    ("ad_time", "广告曝光时间"),
    ("coupon_type", "优惠券类型"),
    ("deal_status", "是否秒杀"),
    ("source", "数据源"),
    ("status", "状态"),
    ("message", "备注"),
]

DB_FIELDS = [key for key, _ in FIELD_COLUMNS]


def app_timezone(name: str | None = None):
    try:
        return ZoneInfo(name or DEFAULT_TIMEZONE)
    except ZoneInfoNotFoundError:
        return timezone(timedelta(hours=8))


def now_local(tz_name: str | None = None) -> datetime:
    return datetime.now(app_timezone(tz_name))


def ensure_storage() -> None:
    for path in (DATA_DIR, EXPORT_DIR, JOB_DIR, DAILY_DIR, CONNECTION_DIR):
        path.mkdir(parents=True, exist_ok=True)
    column_sql = ",\n".join(f'"{field}" TEXT' for field in DB_FIELDS)
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            f"""
            CREATE TABLE IF NOT EXISTS captures (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                owner_id TEXT NOT NULL,
                {column_sql},
                raw_json TEXT
            )
            """
        )
        existing = {row[1] for row in conn.execute("PRAGMA table_info(captures)")}
        if "owner_id" not in existing:
            conn.execute("ALTER TABLE captures ADD COLUMN owner_id TEXT")
        if "raw_json" not in existing:
            conn.execute("ALTER TABLE captures ADD COLUMN raw_json TEXT")
        for field in DB_FIELDS:
            if field not in existing:
                conn.execute(f'ALTER TABLE captures ADD COLUMN "{field}" TEXT')
        conn.execute("CREATE INDEX IF NOT EXISTS idx_capture_owner ON captures(owner_id, id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_capture_lookup ON captures(date, asin, keyword)")


_SCHEDULER_CIPHER: Fernet | None = None
_SCHEDULER_CIPHER_LOCK = threading.Lock()
_SCHEDULER_RUN_LOCK = threading.Lock()


def scheduler_cipher() -> Fernet:
    """Return a stable cipher used to encrypt scheduled credentials at rest."""
    global _SCHEDULER_CIPHER
    with _SCHEDULER_CIPHER_LOCK:
        if _SCHEDULER_CIPHER is not None:
            return _SCHEDULER_CIPHER
        configured = (os.getenv("SCHEDULER_SECRET_KEY") or "").strip()
        if configured:
            key = configured.encode("utf-8")
        else:
            SCHEDULER_KEY_PATH.parent.mkdir(parents=True, exist_ok=True)
            if SCHEDULER_KEY_PATH.exists():
                key = SCHEDULER_KEY_PATH.read_bytes().strip()
            else:
                key = Fernet.generate_key()
                SCHEDULER_KEY_PATH.write_bytes(key)
                try:
                    os.chmod(SCHEDULER_KEY_PATH, 0o600)
                except OSError:
                    pass
        try:
            _SCHEDULER_CIPHER = Fernet(key)
        except (TypeError, ValueError) as exc:
            raise RuntimeError("SCHEDULER_SECRET_KEY 格式无效，请使用 Fernet 密钥") from exc
        return _SCHEDULER_CIPHER


def encrypt_daily_payload(payload: dict[str, Any]) -> str:
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    return scheduler_cipher().encrypt(raw).decode("ascii")


def decrypt_daily_payload(token: str) -> dict[str, Any]:
    try:
        raw = scheduler_cipher().decrypt(str(token or "").encode("ascii"))
        value = json.loads(raw.decode("utf-8"))
    except (InvalidToken, ValueError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise RuntimeError("定时任务凭证无法解密；请重新填写连接并保存定时任务") from exc
    if not isinstance(value, dict):
        raise RuntimeError("定时任务配置格式无效，请重新保存")
    return value


def sanitize_owner_id(value: str) -> str:
    value = (value or "").strip()
    return value if re.fullmatch(r"[A-Za-z0-9_-]{16,100}", value) else ""


def validate_hosted_mcp_url(url: str) -> None:
    parsed = urlparse(url)
    if parsed.scheme.lower() != "https" or not parsed.hostname:
        raise ValueError("Zeabur 模式只允许填写可公网访问的 HTTPS Sorftime MCP URL")
    host = parsed.hostname.lower().rstrip(".")
    if host == "localhost" or host.endswith(".local"):
        raise ValueError("Zeabur 无法访问你电脑的 localhost；本机 MCP/CLI 请使用本机版")
    try:
        address = ipaddress.ip_address(host)
    except ValueError:
        address = None
    if address and (address.is_private or address.is_loopback or address.is_link_local or address.is_reserved):
        raise ValueError("Zeabur 模式不能访问本机或内网 MCP 地址；请填写公网 HTTPS MCP URL")
    if HOSTED_MCP_HOSTS and not any(host == allowed or host.endswith("." + allowed) for allowed in HOSTED_MCP_HOSTS):
        raise ValueError("该 MCP 域名不在 Zeabur 服务允许列表中")


def connection_path(owner_id: str) -> Path:
    safe = sanitize_owner_id(owner_id)
    return CONNECTION_DIR / f"{safe}.json"


def normalize_connection(value: dict[str, Any] | None) -> dict[str, Any]:
    value = value or {}
    mode = str(value.get("mode", "cli_account") or "cli_account").strip()
    if mode not in {"cli_account", "mcp_url", "mcp_stdio"}:
        mode = "cli_account"
    return {
        "mode": mode,
        "mcp_url": str(value.get("mcp_url", "") or "").strip(),
        "mcp_token": str(value.get("mcp_token", "") or "").strip(),
        "cli_account_sk": str(value.get("cli_account_sk", "") or "").strip(),
        "cli_command": str(value.get("cli_command", "") or "").strip(),
        "cli_cwd": str(value.get("cli_cwd", "") or "").strip(),
    }


def connection_has_value(connection: dict[str, Any]) -> bool:
    mode = connection.get("mode")
    if mode == "mcp_url":
        return bool(connection.get("mcp_url"))
    if mode == "cli_account":
        return bool(connection.get("cli_account_sk"))
    return bool(connection.get("cli_command"))


def load_local_connection(owner_id: str) -> dict[str, Any] | None:
    if HOSTED_MODE:
        return None
    path = connection_path(owner_id)
    data = read_json(path) if path.exists() else None
    return normalize_connection(data) if data else None


def save_local_connection(owner_id: str, connection: dict[str, Any]) -> None:
    if HOSTED_MODE:
        raise ValueError("Zeabur 模式不会保存 Sorftime URL、Key 或 CLI；连接仅用于当前请求")
    if not sanitize_owner_id(owner_id):
        raise ValueError("浏览器任务标识无效，无法在本机保存 Sorftime 连接")
    normalized = normalize_connection(connection)
    if not connection_has_value(normalized):
        raise ValueError("Sorftime 连接信息为空")
    path = connection_path(owner_id)
    write_json_atomic(path, normalized)
    try:
        os.chmod(path, 0o600)
    except OSError:
        pass


def clear_local_connection(owner_id: str) -> None:
    if HOSTED_MODE:
        return
    path = connection_path(owner_id)
    if path.exists():
        path.unlink()


def mask_connection(connection: dict[str, Any] | None) -> str:
    if not connection:
        return "未保存"
    if connection.get("mode") == "mcp_url":
        url = str(connection.get("mcp_url", ""))
        base = re.sub(r"([?&](?:key|token|api_key)=)[^&]+", r"\1••••••", url, flags=re.I)
        return base[:120]
    if connection.get("mode") == "cli_account":
        value = str(connection.get("cli_account_sk", ""))
        return f"CLI Account-SK ·••••{value[-4:]}" if value else "CLI Account-SK"
    command = str(connection.get("cli_command", ""))
    return (command[:90] + "…") if len(command) > 90 else command


def default_download_dir() -> Path:
    downloads = Path.home() / "Downloads"
    base = downloads if downloads.exists() else Path.home()
    return base / "AmazonKeywordTracker"


def resolve_download_dir(value: str | None) -> Path:
    raw = os.path.expandvars(str(value or "").strip())
    path = Path(raw).expanduser() if raw else default_download_dir()
    if not path.is_absolute():
        raise ValueError("Excel 保存目录必须填写本机绝对路径")
    path.mkdir(parents=True, exist_ok=True)
    test_path = path / ".keyword-tracker-write-test"
    try:
        test_path.write_text("ok", encoding="utf-8")
        test_path.unlink()
    except OSError as exc:
        raise ValueError(f"Excel 保存目录不可写：{path}") from exc
    return path


def parse_text_items(value: str, uppercase: bool = False) -> list[str]:
    items: list[str] = []
    for raw in re.split(r"[\n,;，；]+", value or ""):
        item = raw.strip()
        if item:
            items.append(item.upper() if uppercase else item)
    return dedupe(items)


def dedupe(items: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        key = item.casefold()
        if key not in seen:
            seen.add(key)
            result.append(item)
    return result


def read_items_from_upload(field: Any | None, uppercase: bool = False) -> list[str]:
    if field is None or not getattr(field, "filename", ""):
        return []
    data = field.file.read()
    filename = field.filename.lower()
    if filename.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        result: list[str] = []
        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            value = "" if not row or row[0] is None else str(row[0]).strip()
            if not value:
                continue
            if index == 1 and value.casefold() in {"asin", "keyword", "关键词"}:
                continue
            result.append(value.upper() if uppercase else value)
        return dedupe(result)
    text = decode_text(data)
    return parse_text_items(text, uppercase)


def decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def traffic_share_number(value: Any) -> float:
    text = str(value or "").replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return -1.0
    number = float(match.group())
    return number / 100 if "%" in text else number


def sort_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    asin_order: dict[str, int] = {}
    for record in records:
        asin_order.setdefault(str(record.get("asin", "")), len(asin_order))
    return sorted(
        records,
        key=lambda record: (
            asin_order.get(str(record.get("asin", "")), 999999),
            -traffic_share_number(record.get("traffic_share")),
            str(record.get("keyword", "")).casefold(),
        ),
    )


def make_workbook(records: list[dict[str, Any]], stats: dict[str, Any]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "关键词监控结果"
    sheet.append([label for _, label in FIELD_COLUMNS])
    for record in sort_records(records):
        sheet.append([record.get(key, "") for key, _ in FIELD_COLUMNS])

    header_fill = PatternFill("solid", fgColor="176B87")
    header_font = Font(name=EXCEL_FONT, size=11, bold=True, color="FFFFFF")
    body_font = Font(name=EXCEL_FONT, size=10)
    align = Alignment(horizontal="left", vertical="center", wrap_text=False)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = align

    widths = [12, 15, 28, 13, 13, 12, 11, 11, 11, 16, 11, 11, 11, 13, 9, 11, 40,
              20, 10, 20, 20, 13, 12, 14, 12, 36]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    summary = workbook.create_sheet("任务汇总")
    summary_rows = [
        ["指标", "数值"],
        ["记录数", len(records)],
        ["Sorftime接口调用总次数", stats.get("mcp_calls", 0)],
        ["运行时间（秒）", stats.get("elapsed_seconds", 0)],
        ["生成时间", now_local().isoformat(timespec="seconds")],
        [],
        ["Sorftime接口", "调用次数", "耗时（秒）"],
    ]
    summary.append(summary_rows[0])
    for row in summary_rows[1:]:
        summary.append(row)
    calls = stats.get("tool_calls", {}) or {}
    seconds = stats.get("tool_seconds", {}) or {}
    for name, count in calls.items():
        summary.append([name, count, seconds.get(name, 0)])
    for row in summary.iter_rows():
        for cell in row:
            cell.font = body_font
            cell.alignment = align
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    for cell in summary[7]:
        cell.fill = PatternFill("solid", fgColor="64CCC5")
        cell.font = Font(name=EXCEL_FONT, bold=True, color="0B2447")
    summary.column_dimensions["A"].width = 36
    summary.column_dimensions["B"].width = 20
    summary.column_dimensions["C"].width = 20

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def make_template(kind: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    label = "关键词" if kind == "keywords" else "ASIN"
    sheet.title = "keywords" if kind == "keywords" else "asins"
    sheet.append([label])
    sheet.column_dimensions["A"].width = 30
    sheet["A1"].fill = PatternFill("solid", fgColor="176B87")
    sheet["A1"].font = Font(name=EXCEL_FONT, bold=True, color="FFFFFF")
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def save_records(records: list[dict[str, Any]]) -> None:
    if not records:
        return
    columns = ["owner_id", *DB_FIELDS, "raw_json"]
    placeholders = ",".join("?" for _ in columns)
    names = ",".join(f'"{column}"' for column in columns)
    rows = []
    for record in records:
        rows.append(
            [
                str(record.get("owner_id", "")),
                *[str(record.get(field, "")) for field in DB_FIELDS],
                json.dumps(record.get("raw", {}), ensure_ascii=False),
            ]
        )
    with sqlite3.connect(DB_PATH) as conn:
        conn.executemany(
            f"INSERT INTO captures ({names}) VALUES ({placeholders})",
            rows,
        )


def latest_history(owner_id: str, limit: int = 200) -> list[dict[str, Any]]:
    owner_id = sanitize_owner_id(owner_id)
    if not owner_id:
        return []
    selected = ",".join(f'"{field}"' for field in DB_FIELDS)
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            f"SELECT {selected} FROM captures WHERE owner_id=? ORDER BY id DESC LIMIT ?",
            (owner_id, limit),
        ).fetchall()
    return [dict(row) for row in rows]


class UploadedField:
    def __init__(self, filename: str, data: bytes) -> None:
        self.filename = filename
        self.file = BytesIO(data)


class SimpleForm:
    def __init__(self) -> None:
        self.values: dict[str, list[str]] = {}
        self.files: dict[str, list[UploadedField]] = {}

    def add_value(self, name: str, value: str) -> None:
        self.values.setdefault(name, []).append(value)

    def add_file(self, name: str, field: UploadedField) -> None:
        self.files.setdefault(name, []).append(field)

    def getfirst(self, name: str, default: str = "") -> str:
        values = self.values.get(name)
        return values[0] if values else default

    def __contains__(self, name: str) -> bool:
        return name in self.values or name in self.files

    def __getitem__(self, name: str) -> Any:
        if name in self.files:
            values = self.files[name]
            return values[0] if len(values) == 1 else values
        values = self.values.get(name, [])
        return values[0] if len(values) == 1 else values


def form_from_request(handler: BaseHTTPRequestHandler) -> SimpleForm:
    content_type = handler.headers.get("Content-Type", "")
    length = int(handler.headers.get("Content-Length", "0") or 0)
    body = handler.rfile.read(length)
    form = SimpleForm()
    if "application/x-www-form-urlencoded" in content_type:
        for key, values in parse_qs(body.decode("utf-8", errors="replace"), keep_blank_values=True).items():
            for value in values:
                form.add_value(key, value)
        return form
    if "multipart/form-data" in content_type:
        message = BytesParser(policy=email_default_policy).parsebytes(
            f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode() + body
        )
        for part in message.iter_parts():
            name = part.get_param("name", header="content-disposition")
            if not name:
                continue
            payload = part.get_payload(decode=True) or b""
            filename = part.get_filename()
            if filename:
                form.add_file(name, UploadedField(filename, payload))
            else:
                charset = part.get_content_charset() or "utf-8"
                form.add_value(name, payload.decode(charset, errors="replace"))
        return form
    if "application/json" in content_type and body:
        data = json.loads(body.decode("utf-8"))
        for key, value in data.items():
            form.add_value(key, "" if value is None else str(value))
    return form


def first_form_field(form: SimpleForm, *names: str) -> Any | None:
    for name in names:
        if name in form:
            field = form[name]
            return field[0] if isinstance(field, list) and field else field
    return None


def parse_capture_form(form: SimpleForm) -> dict[str, Any]:
    asins = parse_text_items(form.getfirst("asins_text") or form.getfirst("asinText"), True)
    keywords = parse_text_items(form.getfirst("keywords_text") or form.getfirst("keywordText"))
    asins += read_items_from_upload(first_form_field(form, "asins_file", "asinFile"), True)
    keywords += read_items_from_upload(first_form_field(form, "keywords_file", "keywordFile"))
    output_mode = form.getfirst("outputMode") or form.getfirst("delivery") or "excel"
    if output_mode == "feishu":
        output_mode = "lark"
    return {
        "asins": dedupe(asins),
        "keywords": dedupe(keywords),
        "marketplace": (form.getfirst("marketplace") or form.getfirst("marketplaceCode") or "US").upper(),
        "owner_id": sanitize_owner_id(form.getfirst("owner_id")),
        "delivery": output_mode,
        "daily_enabled": form.getfirst("daily_enabled") in {"on", "true", "1", "yes"},
        "auto_download": form.getfirst("auto_download", "on") in {"on", "true", "1", "yes"},
        "run_time": form.getfirst("run_time", "09:00") or "09:00",
        "timezone": form.getfirst("timezone", DEFAULT_TIMEZONE) or DEFAULT_TIMEZONE,
        "download_dir": form.getfirst("download_dir").strip(),
        "remember_connection": form.getfirst("remember_connection", "on") in {"on", "true", "1", "yes"},
        "connection": {
            "mode": form.getfirst("sorftime_mode", "cli_account") or "cli_account",
            "mcp_url": form.getfirst("sorftime_mcp_url").strip(),
            "mcp_token": form.getfirst("sorftime_mcp_token").strip(),
            "cli_account_sk": form.getfirst("sorftime_cli_account_sk").strip(),
            "cli_command": form.getfirst("sorftime_cli_command").strip(),
            "cli_cwd": form.getfirst("sorftime_cli_cwd").strip(),
        },
        "lark": {
            "feishu_app_id": form.getfirst("feishu_app_id").strip(),
            "feishu_app_secret": form.getfirst("feishu_app_secret").strip(),
            "base_url": form.getfirst("base_url").strip(),
            "app_token": form.getfirst("app_token").strip(),
            "table_id": form.getfirst("table_id").strip(),
        },
    }


def validate_payload(payload: dict[str, Any], require_inputs: bool = True) -> None:
    if require_inputs and not payload.get("asins"):
        raise ValueError("请至少输入或上传 1 个 ASIN")
    if require_inputs and not payload.get("keywords"):
        raise ValueError("请至少输入或上传 1 个关键词")
    if not re.fullmatch(r"\d{2}:\d{2}", payload.get("run_time", "")):
        raise ValueError("每日运行时间格式应为 HH:MM")
    delivery = str(payload.get("delivery") or "excel").lower()
    if delivery not in {"excel", "lark", "both"}:
        raise ValueError("输出方式只能选择 Excel、飞书或 Excel + 飞书")
    if delivery in {"lark", "both"}:
        lark = payload.get("lark") or {}
        missing = []
        if not str(lark.get("feishu_app_id") or "").strip():
            missing.append("App ID")
        if not str(lark.get("feishu_app_secret") or "").strip():
            missing.append("App Secret")
        if not str(lark.get("base_url") or "").strip():
            missing.append("Base 链接")
        if missing:
            raise ValueError("选择写入飞书后，请填写：" + "、".join(missing))


def resolve_payload_connection(payload: dict[str, Any], *, for_daily: bool = False) -> dict[str, Any]:
    owner_id = payload.get("owner_id", "")
    entered = normalize_connection(payload.get("connection"))

    if HOSTED_MODE:
        if not connection_has_value(entered):
            raise ValueError("请选择 CLI 或 MCP，并填写 Sorftime 连接信息")
        if entered.get("mode") == "mcp_url":
            validate_hosted_mcp_url(entered.get("mcp_url", ""))
        elif entered.get("mode") == "mcp_stdio":
            raise ValueError("Zeabur 不接受任意命令。请选择 CLI Account-SK 或 MCP URL")
        # Hosted daily jobs persist the current connection encrypted in /app/data.
        payload["connection"] = entered
        payload["remember_connection"] = False
        payload["download_dir"] = ""
        payload["_prepared"] = True
        return payload

    saved = load_local_connection(owner_id) if owner_id else None
    if connection_has_value(entered):
        connection = entered
        if payload.get("remember_connection", True):
            save_local_connection(owner_id, connection)
    elif saved and saved.get("mode") == entered.get("mode"):
        connection = saved
    else:
        label = "Sorftime MCP URL" if entered.get("mode") == "mcp_url" else "Sorftime CLI Account-SK"
        raise ValueError(f"请先填写或测试并保存{label}")
    payload["connection"] = connection
    payload["download_dir"] = str(resolve_download_dir(payload.get("download_dir")))
    payload["_prepared"] = True
    return payload


def sanitize_payload_for_disk(payload: dict[str, Any]) -> dict[str, Any]:
    clean = json.loads(json.dumps(payload, ensure_ascii=False))
    clean.pop("_prepared", None)
    lark = clean.get("lark", {})
    lark["feishu_app_secret"] = ""
    connection = normalize_connection(clean.get("connection"))
    clean["connection"] = {
        "mode": connection.get("mode", "cli_account"),
        "mcp_url": "",
        "mcp_token": "",
        "cli_account_sk": "",
        "cli_command": "",
        "cli_cwd": "",
        "saved_locally": not HOSTED_MODE,
    }
    return clean


def job_path(job_id: str) -> Path:
    return JOB_DIR / f"{re.sub(r'[^A-Za-z0-9_-]', '', job_id)}.json"


def daily_path(owner_id: str) -> Path:
    safe = sanitize_owner_id(owner_id)
    return DAILY_DIR / f"{safe}.json"


def write_json_atomic(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temp.replace(path)


def read_json(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None


def append_job_log(job: dict[str, Any], message: str) -> None:
    job.setdefault("logs", []).append(f"{now_local().strftime('%H:%M:%S')} {message}")
    job["logs"] = job["logs"][-150:]


def public_job(job: dict[str, Any]) -> dict[str, Any]:
    result = dict(job)
    result.pop("payload", None)
    result.pop("records", None)
    return result


def public_daily_config(config: dict[str, Any] | None) -> dict[str, Any] | None:
    if not config:
        return None
    return {
        "owner_id": config.get("owner_id", ""),
        "enabled": bool(config.get("enabled")),
        "run_time": config.get("run_time", "09:00"),
        "timezone": config.get("timezone", DEFAULT_TIMEZONE),
        "payload_summary": config.get("payload_summary", {}),
        "last_attempt_date": config.get("last_attempt_date"),
        "last_run_date": config.get("last_run_date"),
        "latest_run_at": config.get("latest_run_at", ""),
        "latest_excel": config.get("latest_excel", ""),
        "latest_stats": config.get("latest_stats", {}),
        "latest_lark": config.get("latest_lark"),
        "latest_records_count": config.get("latest_records_count", 0),
        "last_error": config.get("last_error", ""),
        "storage_note": "Zeabur 请挂载持久化卷到 /app/data；否则重新部署后定时配置会丢失。" if HOSTED_MODE else "",
    }


def save_daily_job(payload: dict[str, Any]) -> dict[str, Any]:
    validate_payload(payload)
    if not payload.get("_prepared"):
        resolve_payload_connection(payload, for_daily=bool(payload.get("daily_enabled")))
    owner_id = payload.get("owner_id", "")
    if not owner_id:
        raise ValueError("浏览器任务标识无效，无法保存每日任务")

    path = daily_path(owner_id)
    existing = read_json(path) or {}
    enabled = bool(payload.get("daily_enabled"))
    schedule_payload = json.loads(json.dumps(payload, ensure_ascii=False))
    schedule_payload.pop("_prepared", None)
    schedule_payload["run_time"] = payload.get("run_time", "09:00") or "09:00"
    schedule_payload["timezone"] = payload.get("timezone", DEFAULT_TIMEZONE) or DEFAULT_TIMEZONE

    current = now_local(schedule_payload["timezone"])
    try:
        scheduled_time = datetime.strptime(schedule_payload["run_time"], "%H:%M").time()
    except ValueError:
        scheduled_time = datetime.strptime("09:00", "%H:%M").time()
    last_attempt_date = existing.get("last_attempt_date")
    # Saving the schedule through a manual run after 09:00 counts as today's attempt,
    # so the scheduler will not duplicate the same batch a few seconds later.
    if enabled and current.time() >= scheduled_time:
        last_attempt_date = current.date().isoformat()

    config = {
        "owner_id": owner_id,
        "enabled": enabled,
        "run_time": schedule_payload["run_time"],
        "timezone": schedule_payload["timezone"],
        "payload_summary": {
            "asin_count": len(schedule_payload.get("asins", [])),
            "keyword_count": len(schedule_payload.get("keywords", [])),
            "marketplace": schedule_payload.get("marketplace", "US"),
            "delivery": schedule_payload.get("delivery", "excel"),
            "connection_mode": (schedule_payload.get("connection") or {}).get("mode", ""),
        },
        "encrypted_payload": encrypt_daily_payload(schedule_payload) if enabled else "",
        "last_attempt_date": last_attempt_date,
        "last_run_date": existing.get("last_run_date"),
        "latest_excel": existing.get("latest_excel", ""),
        "latest_run_at": existing.get("latest_run_at", ""),
        "latest_stats": existing.get("latest_stats", {}),
        "latest_lark": existing.get("latest_lark"),
        "latest_records_count": existing.get("latest_records_count", 0),
        "last_error": existing.get("last_error", ""),
    }
    if not enabled:
        config["last_error"] = ""
    write_json_atomic(path, config)
    return public_daily_config(config) or {}


def run_capture_records(payload: dict[str, Any], progress: Any | None = None):
    validate_payload(payload)
    client = build_sorftime_client(payload.get("connection"))
    records: list[dict[str, Any]] = []
    capture_time = now_local(payload.get("timezone"))
    total = len(payload["asins"]) * len(payload["keywords"])
    done = 0
    try:
        readiness = client.check_ready()
        for asin in payload["asins"]:
            for keyword in payload["keywords"]:
                done += 1
                if progress:
                    progress("before", done, total, asin, keyword, client.stats())
                try:
                    result = client.capture_keyword(asin, keyword, payload["marketplace"])
                except Exception as exc:
                    result = {key: "" for key in [
                        "keyword_rank", "organic_position", "organic_time", "ad_position", "ad_time",
                        "traffic_share", "aba_rank", "search_volume", "price", "coupon_type",
                        "coupon_value", "deal_status", "deal_price", "prime_discount_price",
                        "estimated_sales", "product_rank", "rating", "review_count", "product_url",
                    ]}
                    result.update({"status": "failed", "message": str(exc), "raw": {}})
                record = {
                    "owner_id": payload["owner_id"],
                    "date": capture_time.date().isoformat(),
                    "captured_at": capture_time.isoformat(timespec="seconds"),
                    "marketplace": payload["marketplace"],
                    "asin": asin,
                    "keyword": keyword,
                    **result,
                    "source": client.source_name,
                }
                records.append(record)
                if progress:
                    progress("after", done, total, asin, keyword, client.stats())
        records = sort_records(records)
        save_records(records)
        stats = client.stats()
        stats["connection_check"] = readiness
        return records, stats
    finally:
        client.close()


def write_excel_exports(content: bytes, filename: str, payload: dict[str, Any]) -> tuple[str, str]:
    internal_path = EXPORT_DIR / filename
    internal_path.write_bytes(content)
    local_path = ""
    if payload.get("auto_download", True) and not HOSTED_MODE:
        destination_dir = resolve_download_dir(payload.get("download_dir"))
        destination = destination_dir / filename
        if destination.resolve() != internal_path.resolve():
            shutil.copyfile(internal_path, destination)
        local_path = str(destination)
    return f"/exports/{filename}", local_path


def run_capture_job(job_id: str, payload: dict[str, Any]) -> None:
    job = read_json(job_path(job_id)) or {"id": job_id}
    started = time.perf_counter()
    try:
        def progress(stage: str, done: int, total: int, asin: str, keyword: str, stats: dict[str, Any]) -> None:
            pct = int(5 + ((done - (0 if stage == "after" else 1)) / max(1, total)) * 78)
            job.update({
                "status": "running",
                "done": done if stage == "after" else done - 1,
                "total": total,
                "percent": min(85, max(5, pct)),
                "mcp_calls": stats.get("mcp_calls", 0),
                "elapsed_seconds": round(time.perf_counter() - started, 2),
                "tool_calls": stats.get("tool_calls", {}),
            })
            if stage == "before":
                append_job_log(job, f"抓取 {done}/{total}：{asin} | {keyword}")
            write_json_atomic(job_path(job_id), job)

        records, stats = run_capture_records(payload, progress)
        delivery = payload.get("delivery", "excel")
        job.update({"status": "saving", "percent": 88, "records_count": len(records)})
        if delivery == "excel":
            append_job_log(job, "正在生成 Excel 文件。")
        elif delivery == "lark":
            append_job_log(job, "正在写入飞书 Base。")
        else:
            append_job_log(job, "正在生成 Excel 并写入飞书 Base。")
        write_json_atomic(job_path(job_id), job)

        excel_url = ""
        local_excel_path = ""
        if delivery in {"excel", "both"}:
            output_name = f"amazon-keyword-tracker-{now_local().strftime('%Y%m%d-%H%M%S')}.xlsx"
            excel_url, local_excel_path = write_excel_exports(
                make_workbook(records, stats), output_name, payload
            )

        lark_result = None
        if delivery in {"lark", "both"}:
            lark_result = append_records_to_lark(records, payload.get("lark", {}), FIELD_COLUMNS)
            append_job_log(job, (lark_result or {}).get("message", "飞书写入已完成。"))

        job.update({
            "status": "completed" if not lark_result or lark_result.get("ok") else "completed_with_warning",
            "percent": 100,
            "done": len(payload["asins"]) * len(payload["keywords"]),
            "records_count": len(records),
            "delivery": delivery,
            "excel": excel_url,
            "local_excel_path": local_excel_path,
            "auto_download": bool(payload.get("auto_download", True) and excel_url),
            "mcp_calls": stats.get("mcp_calls", 0),
            "elapsed_seconds": stats.get("elapsed_seconds", round(time.perf_counter() - started, 2)),
            "tool_calls": stats.get("tool_calls", {}),
            "tool_seconds": stats.get("tool_seconds", {}),
            "lark": lark_result,
            "finished_at": now_local().isoformat(timespec="seconds"),
            "records": records,
        })
        append_job_log(job, f"完成：{len(records)} 条，Sorftime 接口 {job['mcp_calls']} 次，用时 {job['elapsed_seconds']} 秒。")
        write_json_atomic(job_path(job_id), job)
    except Exception as exc:
        job.update({
            "status": "failed",
            "error": str(exc),
            "elapsed_seconds": round(time.perf_counter() - started, 2),
            "finished_at": now_local().isoformat(timespec="seconds"),
        })
        append_job_log(job, f"任务失败：{exc}")
        write_json_atomic(job_path(job_id), job)


def create_capture_job(payload: dict[str, Any]) -> dict[str, Any]:
    validate_payload(payload)
    resolve_payload_connection(payload, for_daily=bool(payload.get("daily_enabled")))
    daily_config = None
    if payload.get("owner_id") and "daily_enabled" in payload:
        daily_config = save_daily_job(payload)
    job_id = uuid.uuid4().hex
    total = len(payload["asins"]) * len(payload["keywords"])
    job = {
        "id": job_id,
        "status": "queued",
        "percent": 1,
        "done": 0,
        "total": total,
        "records_count": 0,
        "mcp_calls": 0,
        "elapsed_seconds": 0,
        "tool_calls": {},
        "created_at": now_local().isoformat(timespec="seconds"),
        "logs": [f"{now_local().strftime('%H:%M:%S')} 任务已创建：{len(payload['asins'])} ASIN × {len(payload['keywords'])} 关键词。"],
        "payload": sanitize_payload_for_disk(payload),
        "daily": daily_config,
    }
    write_json_atomic(job_path(job_id), job)
    threading.Thread(target=run_capture_job, args=(job_id, payload), daemon=True).start()
    return public_job(job)


def run_due_daily_jobs_once() -> int:
    """Run every enabled schedule that is due. Returns the number attempted."""
    attempted = 0
    if not _SCHEDULER_RUN_LOCK.acquire(blocking=False):
        return 0
    try:
        for path in DAILY_DIR.glob("*.json"):
            config = read_json(path)
            if not config or not config.get("enabled"):
                continue
            tz_name = config.get("timezone", DEFAULT_TIMEZONE)
            now = now_local(tz_name)
            run_time = config.get("run_time", "09:00")
            try:
                due = datetime.strptime(run_time, "%H:%M").time()
            except ValueError:
                config["last_error"] = f"{now.isoformat(timespec='seconds')} 定时时间格式错误：{run_time}"
                write_json_atomic(path, config)
                continue
            today = now.date().isoformat()
            if now.time() < due or config.get("last_attempt_date") == today:
                continue

            attempted += 1
            # Claim today's run before calling external services, avoiding a retry every minute
            # after a provider error or container thread overlap.
            config["last_attempt_date"] = today
            config["last_error"] = ""
            write_json_atomic(path, config)
            try:
                encrypted = config.get("encrypted_payload", "")
                if not encrypted:
                    raise RuntimeError("定时任务没有保存连接凭证，请在页面重新勾选并开始抓取一次")
                payload = decrypt_daily_payload(encrypted)
                payload["owner_id"] = config.get("owner_id", payload.get("owner_id", ""))
                payload["daily_enabled"] = True
                payload["run_time"] = run_time
                payload["timezone"] = tz_name
                payload["_prepared"] = True

                records, stats = run_capture_records(payload)
                excel_url = ""
                local_excel_path = ""
                if payload.get("delivery", "excel") in {"excel", "both"}:
                    owner_suffix = str(config.get("owner_id", ""))[-8:] or "scheduled"
                    filename = f"daily-keyword-tracker-{owner_suffix}-{now.strftime('%Y%m%d-%H%M%S')}.xlsx"
                    excel_url, local_excel_path = write_excel_exports(make_workbook(records, stats), filename, payload)

                lark_result = None
                if payload.get("delivery") in {"lark", "both"}:
                    lark_result = append_records_to_lark(records, payload.get("lark", {}), FIELD_COLUMNS)

                warning = ""
                if lark_result and not lark_result.get("ok"):
                    warning = lark_result.get("message", "飞书写入失败")
                config.update({
                    "last_run_date": today,
                    "latest_run_at": now.isoformat(timespec="seconds"),
                    "latest_excel": excel_url,
                    "latest_local_excel_path": local_excel_path,
                    "latest_stats": stats,
                    "latest_lark": lark_result,
                    "latest_records_count": len(records),
                    "last_error": warning,
                })
            except Exception as exc:
                config["last_error"] = f"{now.isoformat(timespec='seconds')} {exc}"
            write_json_atomic(path, config)
    finally:
        _SCHEDULER_RUN_LOCK.release()
    return attempted


def scheduler_loop() -> None:
    while True:
        try:
            run_due_daily_jobs_once()
        except Exception as exc:
            (DATA_DIR / "scheduler-error.log").write_text(
                f"{now_local().isoformat()} {exc}\n", encoding="utf-8"
            )
        time.sleep(30)


class Handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self) -> None:
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/":
            return self.serve_file(STATIC_DIR / "index.html")
        if parsed.path.startswith("/static/"):
            return self.serve_file(STATIC_DIR / parsed.path.removeprefix("/static/"))
        if parsed.path == "/api/health":
            return self.send_json({
                "ok": True,
                "time": now_local().isoformat(timespec="seconds"),
                "mode": "hosted" if HOSTED_MODE else "local",
                "hosted": HOSTED_MODE,
                "supports_cli": True,
                "supports_stdio": not HOSTED_MODE,
                "supports_daily": True,
                "daily_time": "09:00",
                "daily_timezone": DEFAULT_TIMEZONE,
                "supports_local_save": not HOSTED_MODE,
            })
        if parsed.path == "/api/history":
            owner_id = parse_qs(parsed.query).get("owner_id", [""])[0]
            return self.send_json({"records": latest_history(owner_id)})
        if parsed.path.startswith("/api/jobs/"):
            parts = parsed.path.strip("/").split("/")
            job_id = parts[2] if len(parts) >= 3 else ""
            job = read_json(job_path(job_id))
            if not job:
                return self.send_json({"ok": False, "error": "任务不存在或服务已重启"}, 404)
            if len(parts) == 4 and parts[3] == "results":
                return self.send_json({"ok": True, "records": job.get("records", [])})
            return self.send_json({"ok": True, "job": public_job(job)})
        if parsed.path == "/api/daily":
            owner_id = parse_qs(parsed.query).get("owner_id", [""])[0]
            config = read_json(daily_path(owner_id)) if sanitize_owner_id(owner_id) else None
            return self.send_json({"ok": True, "job": public_daily_config(config)})
        if parsed.path == "/api/connection":
            owner_id = parse_qs(parsed.query).get("owner_id", [""])[0]
            saved = load_local_connection(owner_id) if sanitize_owner_id(owner_id) else None
            return self.send_json({
                "ok": True,
                "saved": bool(saved),
                "mode": (saved or {}).get("mode", "cli_account"),
                "summary": mask_connection(saved),
                "download_dir": "" if HOSTED_MODE else str(default_download_dir()),
                "local_only": not HOSTED_MODE,
                "hosted": HOSTED_MODE,
            })
        if parsed.path == "/api/template":
            kind = parse_qs(parsed.query).get("kind", ["asins"])[0]
            filename = "keywords-template.xlsx" if kind == "keywords" else "asins-template.xlsx"
            return self.send_bytes(
                make_template(kind),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )
        if parsed.path.startswith("/exports/"):
            return self.serve_file(EXPORT_DIR / parsed.path.removeprefix("/exports/"))
        self.send_error(404)

    def do_POST(self) -> None:
        try:
            if self.path == "/api/connection/test":
                form = form_from_request(self)
                owner_id = sanitize_owner_id(form.getfirst("owner_id"))
                connection = normalize_connection({
                    "mode": form.getfirst("sorftime_mode", "cli_account"),
                    "mcp_url": form.getfirst("sorftime_mcp_url"),
                    "mcp_token": form.getfirst("sorftime_mcp_token"),
                    "cli_account_sk": form.getfirst("sorftime_cli_account_sk"),
                    "cli_command": form.getfirst("sorftime_cli_command"),
                    "cli_cwd": form.getfirst("sorftime_cli_cwd"),
                })
                if HOSTED_MODE:
                    if connection.get("mode") == "mcp_url":
                        validate_hosted_mcp_url(connection.get("mcp_url", ""))
                    elif connection.get("mode") == "mcp_stdio":
                        raise ValueError("Zeabur 不接受任意命令。请选择 CLI Account-SK 或 MCP URL")
                elif not connection_has_value(connection):
                    saved = load_local_connection(owner_id)
                    if saved and saved.get("mode") == connection.get("mode"):
                        connection = saved
                result = test_sorftime_connection(connection)
                remember = (not HOSTED_MODE) and form.getfirst("remember_connection", "on") in {"on", "true", "1", "yes"}
                if remember:
                    save_local_connection(owner_id, connection)
                return self.send_json({"ok": True, "connection": result, "saved": remember, "hosted": HOSTED_MODE})
            if self.path == "/api/connection/clear":
                form = form_from_request(self)
                owner_id = sanitize_owner_id(form.getfirst("owner_id"))
                clear_local_connection(owner_id)
                return self.send_json({"ok": True})
            if self.path == "/api/jobs":
                payload = parse_capture_form(form_from_request(self))
                return self.send_json({"ok": True, "job": create_capture_job(payload)})
            if self.path == "/api/daily":
                payload = parse_capture_form(form_from_request(self))
                validate_payload(payload)
                return self.send_json({"ok": True, "job": save_daily_job(payload)})
        except ValueError as exc:
            return self.send_json({"ok": False, "error": str(exc)}, 400)
        except Exception as exc:
            return self.send_json({"ok": False, "error": str(exc)}, 500)
        self.send_error(404)

    def serve_file(self, path: Path) -> None:
        try:
            resolved = path.resolve()
            allowed_roots = (BASE_DIR.resolve(), EXPORT_DIR.resolve())
            allowed = any(resolved == root or root in resolved.parents for root in allowed_roots)
            if not resolved.is_file() or not allowed:
                return self.send_error(404)
            content_type = mimetypes.guess_type(resolved.name)[0] or "application/octet-stream"
            headers = {}
            if resolved.suffix == ".xlsx":
                headers["Content-Disposition"] = f'attachment; filename="{resolved.name}"'
            self.send_bytes(resolved.read_bytes(), content_type, headers=headers)
        except OSError:
            self.send_error(404)

    def send_json(self, payload: dict[str, Any], status: int = 200) -> None:
        self.send_bytes(
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8",
            status=status,
        )

    def send_bytes(
        self,
        body: bytes,
        content_type: str,
        headers: dict[str, str] | None = None,
        status: int = 200,
    ) -> None:
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Cache-Control", "no-store")
        for key, value in (headers or {}).items():
            self.send_header(key, value)
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format: str, *args: Any) -> None:
        return


def main() -> None:
    parser = argparse.ArgumentParser(description="Amazon Keyword Tracker dual mode")
    parser.add_argument("--no-browser", action="store_true", help="不自动打开浏览器")
    args = parser.parse_args()
    ensure_storage()
    if not HOSTED_MODE:
        PID_PATH.write_text(str(os.getpid()), encoding="utf-8")
    threading.Thread(target=scheduler_loop, daemon=True).start()
    server = ThreadingHTTPServer((APP_HOST, APP_PORT), Handler)
    display_host = "127.0.0.1" if APP_HOST == "0.0.0.0" else APP_HOST
    url = f"http://{display_host}:{APP_PORT}"
    print(f"Amazon keyword tracker mode={'hosted' if HOSTED_MODE else 'local'} listening on {APP_HOST}:{APP_PORT}", flush=True)
    if not HOSTED_MODE and not args.no_browser:
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    finally:
        if not HOSTED_MODE:
            try:
                PID_PATH.unlink()
            except OSError:
                pass


if __name__ == "__main__":
    main()
