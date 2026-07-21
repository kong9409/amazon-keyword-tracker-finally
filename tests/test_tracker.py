from __future__ import annotations

import io
import json
import os
import tempfile
import threading
import time
import unittest
from unittest.mock import patch
from pathlib import Path

from openpyxl import load_workbook

import app
import lark_writer
from sorftime_adapter import (
    SorftimeCliClient, SorftimeMcpClient, adapt_schema_value, adapt_tool_arguments, build_tool_name_map,
    find_value, parse_tool_result, validate_sorftime_payload,
    PRICE_KEYS, SALES_KEYS, RATING_KEYS, REVIEW_KEYS,
)


class FakeSorftimeClient(SorftimeMcpClient):
    def __init__(self) -> None:
        super().__init__("https://example.invalid/mcp")
        self.seen: list[tuple[str, dict]] = []

    def _ensure_initialized(self) -> None:
        self._initialized = True

    def call_tool(self, name: str, arguments: dict):
        started = time.perf_counter()
        self.seen.append((name, dict(arguments)))
        with self._lock:
            self._mcp_calls += 1
            self._tool_calls[name] += 1
            self._tool_seconds[name] += time.perf_counter() - started

        if name == "product_traffic_terms":
            return {"data": [{"keyword": "under bed storage", "trafficShare": "12.5%"}]}
        if name == "keyword_detail":
            return {"data": {"keyword": arguments["keyword"], "searchFrequencyRank": 321, "searchVolume": 12345}}
        if name == "keyword_search_results" and arguments["positionType"] == 0:
            return {"data": [
                {"asin": "B000000001", "position": 4},
                {"asin": "B000000002", "position": 9},
            ]}
        if name == "keyword_search_results" and arguments["positionType"] == 2:
            return {"data": []}
        if name == "product_detail":
            asin = arguments["asin"]
            suffix = int(asin[-1])
            return {"data": {
                "asin": asin,
                "price": 29.99 + suffix,
                "coupon": "$5 off",
                "dealPrice": 25.99,
                "primePrice": 27.99,
                "monthlySales": 600 + suffix,
                "BSR": 1000 + suffix,
                "rating": 4.6,
                "reviewCount": 800 + suffix,
            }}
        raise AssertionError(f"Unexpected tool call: {name} {arguments}")


class TrackerTests(unittest.TestCase):


    def test_mcp_structured_content_and_fenced_json_are_parsed(self) -> None:
        structured = {"structuredContent": {"Code": 0, "Data": {"ListingPriceAmount": 39.99, "SalesVolumeOfMonth": 888}}}
        parsed = parse_tool_result(structured)
        self.assertEqual(find_value(parsed, PRICE_KEYS), 39.99)
        self.assertEqual(find_value(parsed, SALES_KEYS), 888)

        fenced = {"content": [{"type": "text", "text": "```json\n{\"Code\":0,\"Data\":{\"ReviewScore\":4.7,\"ReviewAmount\":1234}}\n```"}]}
        parsed2 = parse_tool_result(fenced)
        self.assertEqual(find_value(parsed2, RATING_KEYS), 4.7)
        self.assertEqual(find_value(parsed2, REVIEW_KEYS), 1234)

    def test_product_trend_enum_alias_uses_live_schema(self) -> None:
        self.assertEqual(
            adapt_schema_value(
                "productTrendType",
                "Rank",
                {"type": "string", "enum": ["SalesVolume", "SalesAmount", "Price", "Ranking"]},
            ),
            "Ranking",
        )

    def test_mcp_arguments_follow_live_input_schema(self) -> None:
        schema = {
            "type": "object",
            "properties": {"keyword": {}, "amzSite": {}, "pageIndex": {}, "position_type": {}},
            "required": ["keyword", "amzSite"],
            "additionalProperties": False,
        }
        args = adapt_tool_arguments(
            "keyword_search_results",
            {"keyword": "shower door", "keywordSupportSite": "US", "page": 2, "positionType": 0},
            schema,
        )
        self.assertEqual(args, {"keyword": "shower door", "amzSite": "US", "pageIndex": 2, "position_type": 0})

    def test_mcp_arguments_support_live_snake_case_site_names(self) -> None:
        product_schema = {
            "type": "object",
            "properties": {"asin": {}, "amz_site": {}},
            "required": ["asin"],
            "additionalProperties": False,
        }
        product_args = adapt_tool_arguments(
            "product_detail",
            {"asin": "B0DT499THF", "amzSite": "US"},
            product_schema,
        )
        self.assertEqual(product_args, {"asin": "B0DT499THF", "amz_site": "US"})

        keyword_schema = {
            "type": "object",
            "properties": {"keyword": {}, "keyword_support_site": {}},
            "required": ["keyword"],
            "additionalProperties": False,
        }
        keyword_args = adapt_tool_arguments(
            "keyword_detail",
            {"keyword": "shower door", "keywordSupportSite": "US"},
            keyword_schema,
        )
        self.assertEqual(
            keyword_args,
            {"keyword": "shower door", "keyword_support_site": "US"},
        )

        rank_args = adapt_tool_arguments(
            "product_ranking_trend_by_keyword",
            {"asin": "B0DT499THF", "keyword": "shower door", "marketplace": "US"},
            {
                "type": "object",
                "properties": {"asin": {}, "keyword": {}, "amz_site": {}},
                "additionalProperties": False,
            },
        )
        self.assertEqual(rank_args["amz_site"], "US")

    def test_plain_text_parameter_error_is_not_treated_as_empty_data(self) -> None:
        with self.assertRaisesRegex(RuntimeError, "Please specify the site"):
            validate_sorftime_payload(
                "product_detail",
                "Please specify the site to query. See the amz_site parameter description in the method signature.",
            )

    def test_feishu_403_has_actionable_permission_message(self) -> None:
        message = lark_writer.explain_feishu_error(
            RuntimeError("飞书接口 HTTP 403（/open-apis/bitable/v1/apps/x/tables/y/records/batch_create）：Forbidden")
        )
        self.assertIn("bitable:app", message)
        self.assertIn("协作者", message)
        self.assertIn("高级权限", message)

    def test_mcp_tool_mapping_never_routes_amazon_to_tiktok(self) -> None:
        names = [
            "tiktok_product_detail",
            "tiktok_product_trend",
            "temu_product_detail",
            "amazon.product_detail",
            "product_trend",
            "product_traffic_terms",
            "keyword_detail",
            "keyword_search_results",
            "product_report",
            "product_ranking_trend_by_keyword",
        ]
        mapping = build_tool_name_map(names)
        self.assertEqual(mapping["product_detail"], "amazon.product_detail")
        self.assertEqual(mapping["product_trend"], "product_trend")
        self.assertNotIn("tiktok", mapping["product_detail"].lower())
        self.assertNotIn("temu", mapping["product_detail"].lower())


    def test_feishu_fields_403_falls_back_to_direct_record_write(self) -> None:
        calls: list[str] = []

        def fake_request(url, payload=None, headers=None, method="POST"):
            calls.append(url)
            if url.endswith("/auth/v3/tenant_access_token/internal"):
                return {"code": 0, "tenant_access_token": "tenant-token"}
            if "/fields?page_size=100" in url:
                raise RuntimeError(
                    "飞书接口 HTTP 403（/open-apis/bitable/v1/apps/x/tables/y/fields）：Forbidden"
                )
            if url.endswith("/records/batch_create"):
                return {"code": 0, "data": {"records": payload["records"]}}
            raise AssertionError(url)

        with patch("lark_writer.request_json", side_effect=fake_request):
            result = lark_writer.append_records_to_lark(
                [{"date": "2026-07-21", "asin": "B0DT499THF"}],
                {
                    "feishu_app_id": "cli_demo",
                    "feishu_app_secret": "secret",
                    "base_url": "https://example.feishu.cn/base/bascnDemo?table=tblDemo",
                },
                [("date", "日期"), ("asin", "ASIN")],
            )
        self.assertTrue(result["ok"])
        self.assertEqual(result["written"], 1)
        self.assertIn("跳过自动建字段", result["message"])
        self.assertTrue(any(url.endswith("/records/batch_create") for url in calls))

    def test_feishu_text_fields_are_always_serialized_as_strings(self) -> None:
        captured_payloads: list[dict] = []

        def fake_request(url, payload=None, headers=None, method="POST"):
            if url.endswith("/auth/v3/tenant_access_token/internal"):
                return {"code": 0, "tenant_access_token": "tenant-token"}
            if "/fields?page_size=100" in url:
                return {
                    "code": 0,
                    "data": {
                        "items": [
                            {"field_name": "日期", "type": 1},
                            {"field_name": "月销量", "type": 1},
                            {"field_name": "评分", "type": 1},
                        ]
                    },
                }
            if url.endswith("/records/batch_create"):
                captured_payloads.append(payload)
                return {"code": 0, "data": {"records": payload["records"]}}
            raise AssertionError(url)

        with patch("lark_writer.request_json", side_effect=fake_request):
            result = lark_writer.append_records_to_lark(
                [{"date": "2026-07-21", "estimated_sales": 1234, "rating": 4.7}],
                {
                    "feishu_app_id": "cli_demo",
                    "feishu_app_secret": "secret",
                    "base_url": "https://example.feishu.cn/base/bascnDemo?table=tblDemo",
                },
                [("date", "日期"), ("estimated_sales", "月销量"), ("rating", "评分")],
            )
        self.assertTrue(result["ok"])
        fields = captured_payloads[0]["records"][0]["fields"]
        self.assertEqual(fields["月销量"], "1234")
        self.assertEqual(fields["评分"], "4.7")

    def test_feishu_number_field_omits_empty_and_parses_number(self) -> None:
        field_columns = [("estimated_sales", "月销量"), ("product_rank", "大类排名")]
        fields = lark_writer.build_record_fields(
            {"estimated_sales": "1,234", "product_rank": ""},
            field_columns,
            {
                "月销量": {"field_name": "月销量", "type": 2},
                "大类排名": {"field_name": "大类排名", "type": 2},
            },
        )
        self.assertEqual(fields, {"月销量": 1234})

    def test_feishu_base_link_and_missing_field_creation(self) -> None:
        calls: list[tuple[str, str, dict | None]] = []

        def fake_request(url, payload=None, headers=None, method="POST"):
            calls.append((method, url, payload))
            if url.endswith("/auth/v3/tenant_access_token/internal"):
                return {"code": 0, "tenant_access_token": "tenant-token"}
            if "/fields?page_size=100" in url:
                return {"code": 0, "data": {"items": [{"field_name": "日期"}]}}
            if url.endswith("/fields"):
                return {"code": 0, "data": {"field": {"field_name": payload["field_name"]}}}
            if url.endswith("/records/batch_create"):
                return {"code": 0, "data": {"records": payload["records"]}}
            raise AssertionError(url)

        config = {
            "feishu_app_id": "cli_demo",
            "feishu_app_secret": "secret",
            "base_url": "https://example.feishu.cn/base/bascnDemo?table=tblDemo",
        }
        with patch("lark_writer.request_json", side_effect=fake_request):
            result = lark_writer.append_records_to_lark(
                [{"date": "2026-07-21", "asin": "B000000001"}],
                config,
                [("date", "日期"), ("asin", "ASIN")],
            )
        self.assertTrue(result["ok"])
        self.assertEqual(result["written"], 1)
        self.assertEqual(result["created_fields"], ["ASIN"])
        self.assertTrue(any(url.endswith("/records/batch_create") for _, url, _ in calls))

    def test_parse_feishu_wiki_and_base_links(self) -> None:
        base = lark_writer.parse_bitable_url("https://x.feishu.cn/base/bascnABC?table=tbl123")
        wiki = lark_writer.parse_bitable_url("https://x.feishu.cn/wiki/wikcnABC?table=tbl456")
        self.assertEqual(base["app_token"], "bascnABC")
        self.assertEqual(base["table_id"], "tbl123")
        self.assertEqual(wiki["wiki_token"], "wikcnABC")
        self.assertEqual(wiki["table_id"], "tbl456")

    def test_output_mode_requires_feishu_configuration(self) -> None:
        payload = {
            "asins": ["B000000001"],
            "keywords": ["shower door"],
            "run_time": "09:00",
            "delivery": "lark",
            "lark": {"feishu_app_id": "", "feishu_app_secret": "", "base_url": ""},
        }
        with self.assertRaises(ValueError):
            app.validate_payload(payload)
        payload["lark"] = {
            "feishu_app_id": "cli_demo",
            "feishu_app_secret": "secret",
            "base_url": "https://example.feishu.cn/base/bascnExample?table=tblExample",
        }
        app.validate_payload(payload)

    def test_v4_strict_metric_sources_and_product_trend_fallbacks(self) -> None:
        class StrictSourceClient(SorftimeMcpClient):
            def __init__(self) -> None:
                super().__init__("https://example.invalid/mcp")
                self.seen = []

            def _ensure_initialized(self) -> None:
                self._initialized = True

            def call_tool(self, name, arguments):
                self.seen.append((name, dict(arguments)))
                if name == "product_traffic_terms":
                    return {"Data": [{"KeywordName": "shower door", "TrafficRate": "8.8%"}]}
                if name == "keyword_detail":
                    return {"Data": {"keyword": "shower door", "Rank": 456, "monthlySearchVolume": 9999}}
                if name == "keyword_search_results":
                    if arguments["positionType"] == 0:
                        return {"Data": [{"asin": "B000000001", "position": 6}]}
                    return {"Data": []}
                if name == "product_detail":
                    return {"Data": {"price": 49.99, "rating": 4.5, "reviewCount": 321}}
                if name == "product_trend":
                    trend_type = arguments["productTrendType"]
                    if trend_type == "SalesVolume":
                        return {"Data": [{"recordDate": "2026-07", "MonthSalesVolume": 777}]}
                    if trend_type == "Rank":
                        return {}
                    if trend_type == "Ranking":
                        return {"Data": [{"recordDate": "2026-07", "MainCategoryRank": 888}]}
                    if trend_type == "Price":
                        return {}
                    return {}
                if name == "product_ranking_trend_by_keyword":
                    return {}
                raise AssertionError((name, arguments))

        client = StrictSourceClient()
        result = client.capture_keyword("B000000001", "shower door", "US")
        self.assertEqual(result["traffic_share"], "8.8%")
        self.assertEqual(result["aba_rank"], 456)
        self.assertEqual(result["search_volume"], 9999)
        self.assertEqual(result["estimated_sales"], 777)
        self.assertEqual(result["product_rank"], 888)
        calls = [name for name, _ in client.seen]
        self.assertNotIn("product_report", calls)
        self.assertIn("product_traffic_terms", calls)
        self.assertIn("keyword_detail", calls)
        self.assertIn("product_detail", calls)
        self.assertGreaterEqual(calls.count("product_trend"), 3)

    def test_field_fallbacks_and_cache_reduce_calls(self) -> None:
        client = FakeSorftimeClient()
        first = client.capture_keyword("B000000001", "under bed storage", "US")
        second = client.capture_keyword("B000000002", "under bed storage", "US")

        self.assertEqual(first["traffic_share"], "12.5%")
        self.assertEqual(first["aba_rank"], 321)
        self.assertEqual(first["search_volume"], 12345)
        self.assertEqual(first["organic_position"], 4)
        self.assertEqual(second["organic_position"], 9)
        self.assertEqual(first["price"], 30.99)
        self.assertEqual(first["coupon_value"], "$5 off")
        self.assertEqual(first["deal_price"], 25.99)
        self.assertEqual(first["prime_discount_price"], 27.99)
        self.assertEqual(first["estimated_sales"], 601)
        self.assertEqual(first["product_rank"], 1001)
        self.assertEqual(first["rating"], 4.6)
        self.assertEqual(first["review_count"], 801)
        self.assertIn("/dp/B000000001", first["product_url"])

        counts = client.stats()["tool_calls"]
        self.assertEqual(counts["product_traffic_terms"], 2)
        self.assertEqual(counts["keyword_detail"], 1)
        self.assertEqual(counts["keyword_search_results"], 2)
        self.assertEqual(counts["product_detail"], 2)
        self.assertNotIn("product_trend", counts)

    def test_cli_account_mode_uses_fixed_sorftime_commands(self) -> None:
        old_path = os.environ.get("PATH", "")
        with tempfile.TemporaryDirectory() as tmp:
            executable = Path(tmp) / "sorftime"
            executable.write_text(
                "#!/usr/bin/env python3\n"
                "import json, sys\n"
                "args=sys.argv[1:]\n"
                "if not args: sys.exit(2)\n"
                "if args[0] in {'add','use'}:\n"
                "    print(json.dumps({'ok': True}))\n"
                "elif args[0]=='whoami':\n"
                "    print(json.dumps({'account':'keyword-tracker'}))\n"
                "elif args[0]=='api':\n"
                "    endpoint=args[1]\n"
                "    if endpoint=='ASINRequestKeywordv2':\n"
                "        data=[{'keyword':'under bed storage','trafficShare':'12.5%','searchFrequencyRank':321,'searchVolume':12345,'organicPosition':4,'adPosition':2}]\n"
                "    elif endpoint=='ProductRequest':\n"
                "        data={'price':29.99,'coupon':'$5 off','dealPrice':25.99,'primePrice':27.99,'monthlySales':600,'BSR':1000,'rating':4.6,'reviewCount':800}\n"
                "    else:\n"
                "        print(json.dumps({'Code':1,'Message':'unknown endpoint'})); sys.exit(1)\n"
                "    print(json.dumps({'Code':0,'Data':data}))\n"
                "else:\n"
                "    sys.exit(2)\n",
                encoding="utf-8",
            )
            executable.chmod(0o755)
            os.environ["PATH"] = f"{tmp}{os.pathsep}{old_path}"
            try:
                client = SorftimeCliClient("ACCOUNT-SK-SECRET")
                ready = client.check_ready()
                result = client.capture_keyword("B000000001", "under bed storage", "US")
                stats = client.stats()
                client.close()
            finally:
                os.environ["PATH"] = old_path

        self.assertEqual(ready["source"], "sorftime_cli")
        self.assertEqual(result["traffic_share"], "12.5%")
        self.assertEqual(result["organic_position"], 4)
        self.assertEqual(result["ad_position"], 2)
        self.assertEqual(result["price"], 29.99)
        self.assertEqual(result["status"], "ok")
        self.assertEqual(stats["mcp_calls"], 2)
        self.assertEqual(stats["tool_calls"]["ASINRequestKeywordv2"], 1)
        self.assertEqual(stats["tool_calls"]["ProductRequest"], 1)

    def test_excel_has_required_columns_and_call_summary(self) -> None:
        record = {key: "" for key, _ in app.FIELD_COLUMNS}
        record.update({
            "date": "2026-07-21",
            "asin": "B000000001",
            "keyword": "under bed storage",
            "traffic_share": "12.5%",
            "status": "ok",
        })
        stats = {
            "mcp_calls": 7,
            "elapsed_seconds": 3.21,
            "tool_calls": {"keyword_detail": 1, "product_detail": 1},
            "tool_seconds": {"keyword_detail": 0.4, "product_detail": 0.8},
        }
        content = app.make_workbook([record], stats)
        workbook = load_workbook(io.BytesIO(content), data_only=False)
        self.assertEqual(workbook.sheetnames, ["关键词监控结果", "任务汇总"])
        headers = [cell.value for cell in workbook["关键词监控结果"][1]]
        self.assertEqual(headers[:17], [label for _, label in app.FIELD_COLUMNS[:17]])
        summary = workbook["任务汇总"]
        self.assertEqual(summary["B3"].value, 7)
        self.assertEqual(summary["B4"].value, 3.21)

    def test_connection_secrets_are_redacted_from_job_payload(self) -> None:
        connection = {
            "mode": "mcp_url",
            "mcp_url": "https://mcp.sorftime.com/?key=SECRET-IN-URL",
            "mcp_token": "SECRET-TOKEN",
            "cli_account_sk": "SECRET-SK",
        }
        clean = app.sanitize_payload_for_disk({"connection": connection, "lark": {"feishu_app_secret": "SECRET-LARK"}})
        serialized = json.dumps(clean, ensure_ascii=False)
        self.assertNotIn("SECRET", serialized)
        self.assertEqual(clean["connection"]["mcp_url"], "")
        self.assertEqual(clean["connection"]["mcp_token"], "")
        self.assertEqual(clean["connection"]["cli_account_sk"], "")

    def test_hosted_mode_rejects_local_or_insecure_mcp_url(self) -> None:
        app.validate_hosted_mcp_url("https://mcp.sorftime.com/")
        with self.assertRaises(ValueError):
            app.validate_hosted_mcp_url("http://mcp.example.com/path")
        with self.assertRaises(ValueError):
            app.validate_hosted_mcp_url("https://localhost:3000/mcp")
        with self.assertRaises(ValueError):
            app.validate_hosted_mcp_url("https://127.0.0.1/mcp")

    def test_hosted_mode_keeps_excel_for_browser_download(self) -> None:
        original_export = app.EXPORT_DIR
        original_hosted = app.HOSTED_MODE
        try:
            with tempfile.TemporaryDirectory() as internal, tempfile.TemporaryDirectory() as local:
                app.EXPORT_DIR = Path(internal)
                app.HOSTED_MODE = True
                url, local_path = app.write_excel_exports(b"xlsx", "hosted.xlsx", {"auto_download": True, "download_dir": local})
                self.assertEqual(url, "/exports/hosted.xlsx")
                self.assertEqual(local_path, "")
                self.assertEqual((Path(internal) / "hosted.xlsx").read_bytes(), b"xlsx")
                self.assertFalse((Path(local) / "hosted.xlsx").exists())
        finally:
            app.EXPORT_DIR = original_export
            app.HOSTED_MODE = original_hosted

    def test_concurrent_job_updates_are_json_serializable(self) -> None:
        # Smoke-test that ordinary progress payloads remain safe to write from worker threads.
        payload = {"status": "running", "tool_calls": {"ProductRequest": 1}, "logs": ["ok"]}
        errors: list[Exception] = []

        def worker() -> None:
            try:
                json.dumps(payload, ensure_ascii=False)
            except Exception as exc:  # pragma: no cover
                errors.append(exc)

        threads = [threading.Thread(target=worker) for _ in range(4)]
        for thread in threads:
            thread.start()
        for thread in threads:
            thread.join()
        self.assertEqual(errors, [])


if __name__ == "__main__":
    unittest.main()
